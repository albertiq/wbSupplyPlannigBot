import io
import pandas as pd
from aiogram.types import BufferedInputFile
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from services.base import AsyncBaseService


class SupplyReportService(AsyncBaseService):
    def __init__(self, *args, **kwargs):
        pass

    async def __call__(self, planned_supplies: list[dict]) -> BufferedInputFile:
        df = self._prepare_dataframe(planned_supplies)
        return await self.create_excel_file(df)

    @staticmethod
    def _prepare_dataframe(raw_data: list) -> pd.DataFrame:
        """Преобразует сырые данные в DataFrame с обработкой сложных структур"""
        df = pd.DataFrame(raw_data)
        column_rename_map = {
            "vendor_code": "Артикул продавца",
            "barcode": "Баркод",
            "quantity": "Кол-во",
            "warehouse": "Склад",
        }
        df = df.rename(columns=column_rename_map)
        for column in df.columns:
            if df[column].apply(lambda x: isinstance(x, list) and all(isinstance(i, dict) for i in x)).any():
                df[column] = df[column].apply(
                    lambda lst: "\n".join(f"{wh.get('warehouse', '')}:{wh.get('quantity', '')}" for wh in lst)
                    if isinstance(lst, list)
                    else ""
                )

            elif df[column].apply(lambda x: isinstance(x, dict)).any():
                df[column] = df[column].apply(
                    lambda d: "\n".join(f"{k}:{v}" for k, v in d.items()) if isinstance(d, dict) else ""
                )

        total_quantity = df["Кол-во"].sum()
        total_row = {"Баркод": "Итого", "Кол-во": total_quantity}

        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        return df

    @staticmethod
    async def create_excel_file(df: pd.DataFrame) -> BufferedInputFile | None:
        now = datetime.now()
        if df.empty:
            return None

        with io.BytesIO() as buffer:
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                unique_warehouses = [wh for wh in df["Склад"].unique() if not pd.isna(wh)]

                if not unique_warehouses:
                    df.to_excel(writer, sheet_name="Данные", index=False)
                else:
                    for warehouse in unique_warehouses:
                        df_warehouse = df[df["Склад"] == warehouse].copy()

                        if "Склад" in df_warehouse.columns:
                            df_warehouse = df_warehouse.drop(columns=["Склад"])
                        warehouse_total = df_warehouse["Кол-во"].sum()
                        total_row = pd.DataFrame({"Баркод": ["Итого"], "Кол-во": [warehouse_total]})

                        df_warehouse = pd.concat([df_warehouse, total_row], ignore_index=True)

                        sheet_name = str(warehouse)[:31]
                        df_warehouse.to_excel(writer, sheet_name=sheet_name, index=False)

                        ws = writer.sheets[sheet_name]
                        last_row = len(df_warehouse) + 1

                        for col_num in range(1, len(df_warehouse.columns) + 1):
                            cell = ws.cell(row=last_row, column=col_num)
                            cell.font = Font(bold=True)

                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]

                    for col_idx in range(1, worksheet.max_column + 1):
                        max_length = 0
                        col_letter = get_column_letter(col_idx)

                        for row in range(1, worksheet.max_row + 1):
                            cell_value = worksheet.cell(row=row, column=col_idx).value
                            if cell_value is not None:
                                length = len(str(cell_value))
                                if length > max_length:
                                    max_length = length

                        adjusted_width = min(50, max(10, max_length + 2))
                        worksheet.column_dimensions[col_letter].width = adjusted_width

            buffer.seek(0)
            file = BufferedInputFile(buffer.read(), filename=f"supply_planning-{now.strftime('%Y-%m-%d-%H-%M')}.xlsx")
            return file
